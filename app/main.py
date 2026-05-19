"""
Gara Komunale e Diturisë 2026 - Production Server
Architecture: FastAPI + asyncpg (PostgreSQL) + Gunicorn
Handles 250+ concurrent students with connection pooling and answer buffering
"""

import asyncio
import json
import os
import time
import uuid
from collections import defaultdict
from contextlib import asynccontextmanager
from datetime import datetime
from io import BytesIO

import asyncpg
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

def format_duration(seconds):
    if not seconds:
        return "0s"

    seconds = int(seconds)
    minutes = seconds // 60
    secs = seconds % 60

    if minutes == 0:
        return f"{secs}s"

    return f"{minutes}m {secs}s"

# ── CONFIG ────────────────────────────────────────────────────────────────────

DATABASE_URL = os.getenv(
    "DATABASE_URL",
    "postgresql://garauser:garapass@localhost:5432/garadb"
)

# Answer write buffer: batch DB writes every N seconds instead of per-click
ANSWER_BUFFER: dict[str, dict] = defaultdict(dict)  # session_id -> {q_idx: answer}
BUFFER_FLUSH_INTERVAL = 2  # seconds


# ── DB POOL ───────────────────────────────────────────────────────────────────

pool: asyncpg.Pool = None


async def get_pool() -> asyncpg.Pool:
    return pool


@asynccontextmanager
async def lifespan(app: FastAPI):
    global pool
    # Create connection pool: min 5, max 40 connections
    pool = await asyncpg.create_pool(
        DATABASE_URL,
        min_size=5,
        max_size=40,
        command_timeout=30,
        statement_cache_size=0, 
        ssl="require",    # avoid prepared-statement issues with PgBouncer
    )
    await init_db(pool)

    # Start background answer buffer flusher
    task = asyncio.create_task(flush_answers_loop())

    yield

    task.cancel()
    await pool.close()


app = FastAPI(lifespan=lifespan, title="Gara Komunale 2026")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── DB INIT ───────────────────────────────────────────────────────────────────

async def init_db(p: asyncpg.Pool):
    async with p.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS admins (
                id TEXT PRIMARY KEY,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                created_at DOUBLE PRECISION DEFAULT EXTRACT(EPOCH FROM NOW())
            );

            CREATE TABLE IF NOT EXISTS tests (
                id TEXT PRIMARY KEY,
                test_code TEXT UNIQUE NOT NULL,
                subject TEXT NOT NULL,
                grade_levels JSONB NOT NULL,
                questions JSONB NOT NULL,
                duration_seconds INTEGER DEFAULT 3600,
                created_by TEXT,
                created_at DOUBLE PRECISION DEFAULT EXTRACT(EPOCH FROM NOW()),
                active BOOLEAN DEFAULT TRUE
            );

            CREATE TABLE IF NOT EXISTS student_sessions (
                id TEXT PRIMARY KEY,
                test_id TEXT NOT NULL REFERENCES tests(id),
                student_name TEXT NOT NULL,
                student_surname TEXT NOT NULL,
                grade TEXT NOT NULL,
                school TEXT NOT NULL,
                started_at DOUBLE PRECISION,
                submitted_at DOUBLE PRECISION,
                answers JSONB DEFAULT '{}',
                score INTEGER DEFAULT 0,
                max_score INTEGER DEFAULT 100,
                duration_taken DOUBLE PRECISION,
                wrong_questions JSONB DEFAULT '[]',
                correct_questions JSONB DEFAULT '[]',
                status TEXT DEFAULT 'not_started'
            );

            CREATE INDEX IF NOT EXISTS idx_sessions_test_id ON student_sessions(test_id);
            CREATE INDEX IF NOT EXISTS idx_sessions_status   ON student_sessions(status);

            CREATE TABLE IF NOT EXISTS schools (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL
            );
        """)

        # Default admin (password: arionm1234 — store hashed in real prod)
        await conn.execute("""
            INSERT INTO admins (id, username, password_hash)
            VALUES ($1, 'Arion', 'arionm1234')
            ON CONFLICT (username) DO NOTHING
        """, str(uuid.uuid4()))

        # Schools from specification
        schools = [
            ("A. SH.", "Shkolla A. SH."),
            ("B.",     "Shkolla B."),
            ("C.",     "Shkolla C."),
            ("E.",     "Shkolla E."),
            ("D.",     "Shkolla D."),
            ("I. A.",  "Shkolla I. A."),
            ("H. P.",  "Shkolla H. P."),
            ("L. P.",  "Shkolla L. P."),
            ("S",      "Shkolla S"),
        ]
        for sid, sname in schools:
            await conn.execute(
                "INSERT INTO schools (id, name) VALUES ($1, $2) ON CONFLICT (id) DO NOTHING",
                sid, sname
            )


# ── ANSWER BUFFER (batch writes) ─────────────────────────────────────────────

async def flush_answers_loop():
    """Flush buffered answers to DB every BUFFER_FLUSH_INTERVAL seconds.
    This converts 130 individual DB writes/second into one bulk update,
    dramatically reducing connection pressure."""
    while True:
        await asyncio.sleep(BUFFER_FLUSH_INTERVAL)
        if not ANSWER_BUFFER:
            continue
        to_flush = dict(ANSWER_BUFFER)
        ANSWER_BUFFER.clear()
        try:
            async with pool.acquire() as conn:
                for session_id, answers in to_flush.items():
                    await conn.execute(
                        """UPDATE student_sessions
                           SET answers = answers || $1::jsonb
                           WHERE id = $2 AND status = 'in_progress'""",
                        json.dumps(answers), session_id
                    )
        except Exception as e:
            print(f"[buffer flush error] {e}")
            # Re-buffer failed writes
            for sid, ans in to_flush.items():
                ANSWER_BUFFER[sid].update(ans)


# ── SCHEMAS ───────────────────────────────────────────────────────────────────

class AdminLogin(BaseModel):
    username: str
    password: str

class CreateTest(BaseModel):
    subject: str
    grade_levels: list[str]
    questions: list[dict]
    duration_seconds: int = 3600
    admin_id: str | None = None

class StartSession(BaseModel):
    test_code: str
    name: str
    surname: str
    grade: str
    school: str

class SaveAnswer(BaseModel):
    session_id: str
    question_index: int
    answer: int

class SubmitSession(BaseModel):
    session_id: str
    answers: dict | None = None


# ── AUTH ──────────────────────────────────────────────────────────────────────

@app.post("/api/admin/login")
async def admin_login(data: AdminLogin):
    async with pool.acquire() as conn:
        admin = await conn.fetchrow(
            "SELECT * FROM admins WHERE username=$1 AND password_hash=$2",
            data.username, data.password
        )
    if not admin:
        raise HTTPException(401, "Kredencialet janë të gabuara")
    return {"success": True, "adminId": admin["id"], "username": admin["username"]}


# ── SCHOOLS ───────────────────────────────────────────────────────────────────

@app.get("/api/schools")
async def get_schools():
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM schools ORDER BY name")
    return [dict(r) for r in rows]


# ── TESTS ─────────────────────────────────────────────────────────────────────

@app.post("/api/tests")
async def create_test(data: CreateTest):
    tid = str(uuid.uuid4())
    async with pool.acquire() as conn:
        # Generate unique 6-char code
        while True:
            code = str(uuid.uuid4())[:6].upper()
            exists = await conn.fetchval("SELECT id FROM tests WHERE test_code=$1", code)
            if not exists:
                break
        await conn.execute("""
            INSERT INTO tests (id, test_code, subject, grade_levels, questions, duration_seconds, created_by)
            VALUES ($1,$2,$3,$4,$5,$6,$7)
        """, tid, code, data.subject,
            json.dumps(data.grade_levels),
            json.dumps(data.questions),
            data.duration_seconds, data.admin_id)
    return {"success": True, "test_id": tid, "test_code": code}


@app.get("/api/tests")
async def get_tests():
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id,test_code,subject,grade_levels,created_at,active FROM tests ORDER BY created_at DESC"
        )
    result = []
    for r in rows:
        d = dict(r)
        d["grade_levels"] = json.loads(d["grade_levels"]) if isinstance(d["grade_levels"], str) else d["grade_levels"]
        result.append(d)
    return result


@app.get("/api/tests/{test_id}")
async def get_test(test_id: str):
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT * FROM tests WHERE id=$1 OR test_code=$1",
            test_id
        )
    if not row:
        raise HTTPException(404, "Testi nuk u gjet")
    d = dict(row)
    d["questions"]    = json.loads(d["questions"]) if isinstance(d["questions"], str) else d["questions"]
    d["grade_levels"] = json.loads(d["grade_levels"]) if isinstance(d["grade_levels"], str) else d["grade_levels"]
    return d


@app.delete("/api/tests/{test_id}")
async def delete_test(test_id: str):
    async with pool.acquire() as conn:
        await conn.execute("UPDATE tests SET active=FALSE WHERE id=$1", test_id)
    return {"success": True}


# ── STUDENT SESSION ───────────────────────────────────────────────────────────

@app.post("/api/session/start")
async def start_session(data: StartSession):
    async with pool.acquire() as conn:
        test = await conn.fetchrow(
            "SELECT * FROM tests WHERE (id=$1 OR test_code=$1) AND active=TRUE",
            data.test_code
        )
        if not test:
            raise HTTPException(404, "Kodi i testit është i gabuar ose testi nuk është aktiv")

        sid = str(uuid.uuid4())
        now = time.time()
        questions = json.loads(test["questions"]) if isinstance(test["questions"], str) else test["questions"]
        max_score = len(questions) * (10 if len(questions) == 10 else 5)

        await conn.execute("""
            INSERT INTO student_sessions
              (id, test_id, student_name, student_surname, grade, school, started_at, status, max_score)
            VALUES ($1,$2,$3,$4,$5,$6,$7,'in_progress',$8)
        """, sid, test["id"], data.name, data.surname, data.grade, data.school, now, max_score)

    safe_questions = [
        {"index": i, "question": q["question"], "options": q["options"]}
        for i, q in enumerate(questions)
    ]

    return {
        "session_id": sid,
        "test_id": test["id"],
        "subject": test["subject"],
        "duration_seconds": test["duration_seconds"],
        "questions": safe_questions,
        "started_at": now,
    }


@app.post("/api/session/save-answer")
async def save_answer(data: SaveAnswer):
    """Buffer answer in memory — flushed to DB every 2 seconds.
    Returns immediately without hitting the DB, keeping response time <5ms
    even with 250 concurrent students clicking answers."""
    ANSWER_BUFFER[data.session_id][str(data.question_index)] = data.answer
    return {"success": True, "buffered": True}


@app.post("/api/session/submit")
async def submit_session(data: SubmitSession):
    # Flush any buffered answers for this session immediately
    buffered = ANSWER_BUFFER.pop(data.session_id, {})

    async with pool.acquire() as conn:
        session = await conn.fetchrow(
            "SELECT * FROM student_sessions WHERE id=$1", data.session_id
        )
        if not session:
            raise HTTPException(404, "Sesioni nuk u gjet")

        if session["status"] == "completed":
            d = dict(session)
            d["wrong_questions"]   = json.loads(d["wrong_questions"]) if isinstance(d["wrong_questions"], str) else d["wrong_questions"]
            d["correct_questions"] = json.loads(d["correct_questions"]) if isinstance(d["correct_questions"], str) else d["correct_questions"]
            d["answers"]           = json.loads(d["answers"]) if isinstance(d["answers"], str) else d["answers"]
            return d

        test = await conn.fetchrow("SELECT * FROM tests WHERE id=$1", session["test_id"])
        questions = json.loads(test["questions"]) if isinstance(test["questions"], str) else test["questions"]

        # Merge buffered + saved + submitted answers
        saved = json.loads(session["answers"]) if isinstance(session["answers"], str) else (session["answers"] or {})
        saved.update(buffered)
        if data.answers:
            saved.update({str(k): v for k, v in data.answers.items()})

        now = time.time()
        duration = now - session["started_at"]
        points_per_q = 10 if len(questions) == 10 else 5
        score, wrong, correct = 0, [], []

        for i, q in enumerate(questions):
            ans = saved.get(str(i))
            if ans == q["correct"]:
                score += points_per_q
                correct.append(i)
            else:
                wrong.append({
                    "index": i,
                    "question": q["question"],
                    "your_answer": ans,
                    "correct_answer": q["correct"],
                    "options": q["options"],
                })

        await conn.execute("""
            UPDATE student_sessions
            SET status='completed', submitted_at=$1, answers=$2, score=$3,
                duration_taken=$4, wrong_questions=$5, correct_questions=$6
            WHERE id=$7
        """, now, json.dumps(saved), score, duration,
            json.dumps(wrong), json.dumps(correct), data.session_id)

    test_row = dict(test)
    return {
        "session_id":        data.session_id,
        "student_name":      session["student_name"],
        "student_surname":   session["student_surname"],
        "grade":             session["grade"],
        "school":            session["school"],
        "subject":           test_row["subject"],
        "score":             score,
        "max_score":         session["max_score"],
        "duration_taken":    duration,
        "wrong_questions":   wrong,
        "correct_questions": correct,
        "total_questions":   len(questions),
        "submitted_at":      now,
    }


# ── RESULTS ───────────────────────────────────────────────────────────────────

@app.get("/api/results/{test_id}")
async def get_results(test_id: str):
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT ss.*, t.subject FROM student_sessions ss
            JOIN tests t ON ss.test_id = t.id
            WHERE ss.test_id=$1 AND ss.status='completed'
            ORDER BY ss.score DESC, ss.duration_taken ASC
        """, test_id)
    result = []
    for i, r in enumerate(rows):
        d = dict(r)
        d["rank"]              = i + 1
        d["wrong_questions"]   = json.loads(d["wrong_questions"]) if isinstance(d["wrong_questions"], str) else d["wrong_questions"]
        d["correct_questions"] = json.loads(d["correct_questions"]) if isinstance(d["correct_questions"], str) else d["correct_questions"]
        d.pop("answers", None)
        result.append(d)
    return result


@app.get("/api/results/all")
async def get_all_results():
    async with pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT ss.id, ss.student_name, ss.student_surname, ss.grade, ss.school,
                   ss.score, ss.max_score, ss.duration_taken, ss.submitted_at, ss.status,
                   ss.wrong_questions, ss.correct_questions,
                   t.subject, t.test_code
            FROM student_sessions ss
            JOIN tests t ON ss.test_id = t.id
            WHERE ss.status='completed'
            ORDER BY t.subject, ss.score DESC, ss.duration_taken ASC
        """)
    result = []
    for r in rows:
        d = dict(r)
        d["wrong_questions"]   = json.loads(d["wrong_questions"]) if isinstance(d["wrong_questions"], str) else d["wrong_questions"]
        d["correct_questions"] = json.loads(d["correct_questions"]) if isinstance(d["correct_questions"], str) else d["correct_questions"]
        result.append(d)
    return result


# ── EXPORT ────────────────────────────────────────────────────────────────────

@app.get("/api/export/excel/{test_id}")
async def export_excel(test_id: str):
    async with pool.acquire() as conn:
        test = await conn.fetchrow("SELECT * FROM tests WHERE id=$1", test_id)
        if not test:
            raise HTTPException(404, "Testi nuk u gjet")
        sessions = await conn.fetch("""
            SELECT * FROM student_sessions WHERE test_id=$1 AND status='completed'
            ORDER BY score DESC, duration_taken ASC
        """, test_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Rezultatet - {test['subject']}"

    hfill  = PatternFill("solid", fgColor="1a3a5c")
    hfont  = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center")

    headers = ["#", "ID", "Emri", "Mbiemri", "Klasa", "Shkolla",
               "Pikët", "Max", "%", "Kohë(min)", "Gabime", "Saktë", "Data"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment = hfill, hfont, center

    for rank, s in enumerate(sessions, 1):
        dur = format_duration(s["duration_taken"]) if s["duration_taken"] else 0
        pct  = round(s["score"] / s["max_score"] * 100, 1) if s["max_score"] else 0
        wq   = json.loads(s["wrong_questions"]) if isinstance(s["wrong_questions"], str) else s["wrong_questions"]
        cq   = json.loads(s["correct_questions"]) if isinstance(s["correct_questions"], str) else s["correct_questions"]
        date = datetime.fromtimestamp(s["submitted_at"]).strftime("%d/%m/%Y %H:%M") if s["submitted_at"] else ""
        fc   = "e8f5e9" if pct >= 50 else "ffebee"
        for col, val in enumerate([rank, s["id"][:8], s["student_name"], s["student_surname"],
                                    s["grade"], s["school"], s["score"], s["max_score"],
                                    pct, format_duration(s["duration_taken"]), len(wq), len(cq), date], 1):
            cell = ws.cell(row=rank+1, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fc)
            cell.alignment = center

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 4, 30
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Rezultatet_{test['subject'].replace(' ','_')}_{test['test_code']}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


@app.get("/api/export/pdf/{test_id}")
async def export_pdf(test_id: str):
    async with pool.acquire() as conn:
        test = await conn.fetchrow("SELECT * FROM tests WHERE id=$1", test_id)
        if not test:
            raise HTTPException(404, "Testi nuk u gjet")
        sessions = await conn.fetch("""
            SELECT * FROM student_sessions WHERE test_id=$1 AND status='completed'
            ORDER BY score DESC, duration_taken ASC
        """, test_id)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    elems = [
        Paragraph("Gara Komunale e Diturisë 2026", styles["Title"]),
        Paragraph(f"Lënda: {test['subject']} | Kodi: {test['test_code']}", styles["Normal"]),
        Spacer(1, 12),
    ]
    rows = [["#", "ID", "Emri", "Mbiemri", "Klasa", "Shkolla", "Pikët", "%", "Kohë(min)"]]
    for rank, s in enumerate(sessions, 1):
        dur = format_duration(s["duration_taken"]) if s["duration_taken"] else 0
        pct = round(s["score"] / s["max_score"] * 100, 1) if s["max_score"] else 0
        rows.append([rank, s["id"][:8], s["student_name"], s["student_surname"],
                     s["grade"], s["school"], s["score"], f"{pct}%", format_duration(s["duration_taken"])])
    t = Table(rows, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1a3a5c")),
        ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f0f4f8")]),
        ("GRID",       (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    fname = f"Rezultatet_{test['subject'].replace(' ','_')}_{test['test_code']}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ── STATIC ────────────────────────────────────────────────────────────────────

app.mount("/", StaticFiles(directory="public", html=True), name="static")
